# -*- coding: utf-8 -*-
"""PCB BOM 检查工具 - 分析原始 BOM 并输出结构化 JSON 结果。

用法:
    python check_bom.py analyze <xlsx路径>
    python check_bom.py fill <xlsx路径> <行号> <列名> <值>
    python check_bom.py delete <xlsx路径> <行号1> [<行号2> ...]
"""
import sys
import json
import re
from collections import defaultdict, OrderedDict
import openpyxl


# 必须列定义: (显示名, 匹配别名列表)
REQUIRED_COLUMNS = [
    ("品号",       ["品号", "part number", "partnumber", "pn", "料号"]),
    ("物料描述",   ["物料描述", "material description", "description", "描述", "物料名称"]),
    ("位号",       ["designator", "位号", "reference designator", "ref des", "位置"]),
    ("数量",       ["quantity", "数量", "qty"]),
]
# 可选列定义: 有则按层分组，无则整体处理
OPTIONAL_COLUMNS = [
    ("层",         ["layer", "层", "placement"]),
]
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

# 阻值/容值标准化：把物料描述中的数值+单位统一到基础单位(Ω / pF)
# 用于发现"100nF 与 0.1μF"这类因单位写法不同造成的隐式重复
# 电阻：前缀 K/M 可选（"0Ω" 合法）；电容：要求必须有 p/n/μ/u 前缀（避免误匹配型号里的 数字+F）
RES_PATTERN = re.compile(r"(\d+(?:\.\d+)?)\s*([KkM])?Ω")
CAP_PATTERN = re.compile(r"(\d+(?:\.\d+)?)\s*([pnμu])F")


def _norm_resistor(match):
    """电阻值标准化到 Ω。"""
    val = float(match.group(1))
    prefix = match.group(2) or ""
    mult = {"": 1, "K": 1e3, "k": 1e3, "M": 1e6}.get(prefix, 1)
    r = val * mult
    return f"{int(r) if r == int(r) else r}Ω"


def _norm_capacitor(match):
    """电容值标准化到 pF。"""
    val = float(match.group(1))
    prefix = match.group(2)
    mult = {"p": 1, "n": 1e3, "μ": 1e6, "u": 1e6}.get(prefix, 1)
    r = val * mult
    return f"{int(r) if r == int(r) else r}pF"


def normalize_description(desc):
    """返回物料描述的标准化形式（阻值→Ω，容值→pF）。"""
    if not desc:
        return desc
    s = RES_PATTERN.sub(_norm_resistor, str(desc))
    s = CAP_PATTERN.sub(_norm_capacitor, s)
    return s


def _norm(val):
    """标准化表头文本用于比较。"""
    if val is None:
        return ""
    return str(val).strip().lower().replace(" ", "")


def _is_blank(val):
    """判断单元格是否为空白。"""
    if val is None:
        return True
    s = str(val).strip()
    return s == ""


def _match_column(header_val):
    """返回标准列名，匹配不到则返回 None。"""
    nv = _norm(header_val)
    for std_name, aliases in ALL_COLUMNS:
        if nv in [_norm(a) for a in aliases]:
            return std_name
    return None


def analyze(file_path):
    """分析 BOM 文件，输出 JSON。

    列检查：品号/物料描述/位号/数量 为必须，层 为可选。
    若存在"层"列，空白/重复/隐式重复检查按层(Top/Bottom)分别进行。
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers_raw = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_count = len(headers_raw)

    # --- 检查 1: 列结构 ---
    # 建立 列索引 → 标准名 的映射
    col_map = {}
    for idx, h in enumerate(headers_raw):
        std = _match_column(h)
        if std:
            col_map[std] = idx + 1  # 1-based

    # 检查必须列是否齐全
    required_missing = [name for name, _ in REQUIRED_COLUMNS if name not in col_map]
    has_layer = "层" in col_map

    result = {
        "file": file_path,
        "sheet": ws.title,
        "total_rows": ws.max_row,
        "column_check": {
            "required_columns": [name for name, _ in REQUIRED_COLUMNS],
            "optional_columns": [name for name, _ in OPTIONAL_COLUMNS],
            "actual_count": col_count,
            "actual_headers": headers_raw,
            "required_missing": required_missing,
            "has_layer": has_layer,
            "passed": len(required_missing) == 0,
        },
    }

    # 必须列缺失 → 输出并退出
    if required_missing:
        result["blanks"] = []
        result["duplicates"] = []
        result["implicit_duplicates"] = []
        result["has_layer"] = False
        result["summary"] = {
            "column_check_passed": False,
            "blank_count": 0,
            "duplicate_count": 0,
            "implicit_duplicate_count": 0,
            "all_passed": False,
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    def _cell(row, std_name):
        col = col_map.get(std_name)
        if col is None:
            return None
        return ws.cell(row=row, column=col).value

    result["has_layer"] = has_layer

    # 收集所有层值（如果存在层列）
    layers = []
    if has_layer:
        layer_set = OrderedDict()
        for r in range(2, ws.max_row + 1):
            lv = _cell(r, "层")
            if not _is_blank(lv):
                lv_str = str(lv).strip()
                if lv_str not in layer_set:
                    layer_set[lv_str] = True
        layers = list(layer_set.keys())
    result["layers"] = layers

    # 确定要检查空白的列：必须列 + （若有层列则加层）
    cols_to_check = [name for name, _ in REQUIRED_COLUMNS]
    if has_layer:
        cols_to_check.append("层")

    # --- 检查 2: 空白单元格 ---
    blanks = []
    for r in range(2, ws.max_row + 1):
        for std_name in cols_to_check:
            val = _cell(r, std_name)
            if _is_blank(val):
                blanks.append({
                    "row": r,
                    "column": std_name,
                    "designator": _cell(r, "位号"),
                    "part_number": _cell(r, "品号"),
                    "description": _cell(r, "物料描述"),
                    "layer": _cell(r, "层") if has_layer else None,
                    "current_value": val,
                })
    result["blanks"] = blanks

    # --- 检查 3: 品号重复 ---
    # 有层列时，按 (品号, 层) 分组；无层列时，按 品号 分组
    pn_rows = defaultdict(list)
    pn_order = OrderedDict()
    for r in range(2, ws.max_row + 1):
        pn = _cell(r, "品号")
        if _is_blank(pn):
            continue
        pn_str = str(pn).strip()
        layer_val = _cell(r, "层") if has_layer else None
        layer_str = str(layer_val).strip() if not _is_blank(layer_val) else ""
        group_key = (pn_str, layer_str) if has_layer else pn_str

        pn_rows[group_key].append({
            "row": r,
            "designator": _cell(r, "位号"),
            "description": _cell(r, "物料描述"),
            "quantity": _cell(r, "数量"),
            "layer": layer_val if has_layer else None,
            "part_number": pn_str,
        })
        if group_key not in pn_order:
            pn_order[group_key] = True

    duplicates = []
    for key in pn_order:
        rows = pn_rows[key]
        if len(rows) > 1:
            duplicates.append({"part_number": rows[0]["part_number"], "occurrences": rows})
    result["duplicates"] = duplicates

    # --- 检查 4: 物料描述隐式重复（单位不同但实际值相同，如 100nF 与 0.1μF）---
    # 有层列时，按 (标准化描述, 层) 分组
    desc_groups = defaultdict(list)
    desc_order = OrderedDict()
    for r in range(2, ws.max_row + 1):
        desc = _cell(r, "物料描述")
        if _is_blank(desc):
            continue
        desc_str = str(desc).strip()
        norm_desc = normalize_description(desc_str)
        layer_val = _cell(r, "层") if has_layer else None
        layer_str = str(layer_val).strip() if not _is_blank(layer_val) else ""
        group_key = (norm_desc, layer_str) if has_layer else norm_desc

        desc_groups[group_key].append({
            "row": r,
            "raw_description": desc_str,
            "normalized_description": norm_desc,
            "part_number": _cell(r, "品号"),
            "designator": _cell(r, "位号"),
            "quantity": _cell(r, "数量"),
            "layer": layer_val if has_layer else None,
        })
        if group_key not in desc_order:
            desc_order[group_key] = True

    implicit_duplicates = []
    for key in desc_order:
        items = desc_groups[key]
        raw_variants = {it["raw_description"] for it in items}
        if len(raw_variants) > 1:
            implicit_duplicates.append({
                "normalized_description": items[0]["normalized_description"],
                "occurrences": items,
            })
    result["implicit_duplicates"] = implicit_duplicates

    # 汇总
    result["summary"] = {
        "column_check_passed": True,
        "has_layer": has_layer,
        "layers": layers,
        "blank_count": len(blanks),
        "duplicate_count": len(duplicates),
        "implicit_duplicate_count": len(implicit_duplicates),
        "all_passed": (
            len(blanks) == 0
            and len(duplicates) == 0
            and len(implicit_duplicates) == 0
        ),
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))


def fill_cell(file_path, row, col_name, value):
    """填充单个空白单元格并保存。"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers_raw = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for idx, h in enumerate(headers_raw):
        if _match_column(h) == col_name:
            ws.cell(row=int(row), column=idx + 1).value = value
            wb.save(file_path)
            print(json.dumps({"success": True, "row": row, "column": col_name, "value": value}, ensure_ascii=False))
            return
    print(json.dumps({"success": False, "error": f"列 '{col_name}' 未找到"}, ensure_ascii=False))


def delete_rows(file_path, row_numbers):
    """删除指定行并保存（从大到小删除以保持行号稳定）。"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    rows_sorted = sorted(set(int(r) for r in row_numbers), reverse=True)
    for r in rows_sorted:
        ws.delete_rows(r, 1)
    wb.save(file_path)
    print(json.dumps({"success": True, "deleted_rows": sorted(set(int(r) for r in row_numbers))}, ensure_ascii=False))


# 优先仓库（库存地），按优先级排序
PRIORITY_WAREHOUSES = ["1003", "1004", "1016"]


def extract_parts(file_path):
    """提取去重后的品号列表，聚合相同品号的总需求量。

    跳过没有品号的行（品号为空或值为"虚拟件"）。
    输出 JSON 供第二步库存检查使用。
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers_raw = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    col_map = {}
    for idx, h in enumerate(headers_raw):
        std = _match_column(h)
        if std:
            col_map[std] = idx + 1

    def _cell(row, std_name):
        col = col_map.get(std_name)
        if col is None:
            return None
        return ws.cell(row=row, column=col).value

    # 聚合品号 → 总数量 + 涉及行 + 位号列表
    parts_map = OrderedDict()
    for r in range(2, ws.max_row + 1):
        pn = _cell(r, "品号")
        if _is_blank(pn):
            continue
        pn_str = str(pn).strip()
        if pn_str == "虚拟件":
            continue
        qty_raw = _cell(r, "数量")
        try:
            qty = float(qty_raw) if qty_raw is not None else 0
        except (ValueError, TypeError):
            qty = 0
        desc = _cell(r, "物料描述")
        desig = _cell(r, "位号")

        if pn_str not in parts_map:
            parts_map[pn_str] = {
                "part_number": pn_str,
                "description": str(desc).strip() if desc else "",
                "total_quantity": 0,
                "rows": [],
                "designators": [],
            }
        parts_map[pn_str]["total_quantity"] += qty
        parts_map[pn_str]["rows"].append(r)
        if desig:
            parts_map[pn_str]["designators"].append(str(desig).strip())

    parts_list = []
    for pn, info in parts_map.items():
        total = info["total_quantity"]
        info["total_quantity"] = int(total) if total == int(total) else total
        parts_list.append(info)

    result = {
        "file": file_path,
        "total_unique_parts": len(parts_list),
        "priority_warehouses": PRIORITY_WAREHOUSES,
        "parts": parts_list,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


def gen_picking_list(input_path):
    """根据库存检查结果生成领料单 Excel（多工作表 + 条件标色）。

    输入 JSON 支持两种格式：
    1. 旧格式 — 纯数组 [...]
    2. 新格式 — 对象：
       {
         "items": [...],           # 领料列表
         "deleted_items": [...],   # 第一步删除的行
         "replaced_items": [...]   # 替换对照
       }

    每个 item 可含字段：part_number, description, designators, required_quantity,
    selected_warehouse, selected_batch, available_quantity, status
    """
    with open(input_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    # 兼容旧格式（纯数组）
    if isinstance(raw, list):
        items = raw
        deleted_items = []
        replaced_items = []
    else:
        items = raw.get("items", [])
        deleted_items = raw.get("deleted_items", [])
        replaced_items = raw.get("replaced_items", [])

    # --- 样式定义 ---
    uniform_font = openpyxl.styles.Font(name="等线", size=10)
    header_font = openpyxl.styles.Font(name="等线", size=10, bold=True)
    red_fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = openpyxl.styles.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def _status_text(s):
        return {"ok": "正常", "replacement": "已替换", "no_stock": "无库存"}.get(s, s or "")

    # --- 排序：按仓库分组，仓库内按品号从小到大 ---
    def _sort_key(item):
        ws_val = str(item.get("selected_warehouse", "") or "")
        pn_val = str(item.get("part_number", "") or "")
        # 品号是数字字符串，补零对齐使数值排序正确
        return (ws_val, pn_val.zfill(20))

    items = sorted(items, key=_sort_key)

    def _set_widths(ws, widths):
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    def _style_header(ws, ncols):
        for c in range(1, ncols + 1):
            ws.cell(row=1, column=c).font = header_font

    def _apply_font(ws):
        """统一全表字体（不覆盖表头加粗和已有的 fill）"""
        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1:
                    continue  # 表头已单独设
                cell.font = uniform_font

    # --- 工作表 1：领料单 ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "领料单"
    headers = ["序号", "品号", "物料描述", "位号", "需求量", "仓库", "批次", "库存量", "状态"]
    ws.append(headers)
    _style_header(ws, len(headers))
    _set_widths(ws, [6, 14, 45, 50, 8, 8, 16, 12, 10])

    for i, item in enumerate(items, 1):
        status_t = _status_text(item.get("status", ""))
        avail = item.get("available_quantity")
        req = item.get("required_quantity", 0)
        # 安全转数字
        try:
            avail_num = float(avail) if avail is not None else None
        except (ValueError, TypeError):
            avail_num = None
        try:
            req_num = float(req)
        except (ValueError, TypeError):
            req_num = 0

        row_data = [
            i,
            item.get("part_number", ""),
            item.get("description", ""),
            item.get("designators", ""),
            req,
            item.get("selected_warehouse", ""),
            item.get("selected_batch", ""),
            avail if avail is not None else "",
            status_t,
        ]
        ws.append(row_data)
        row_idx = i + 1

        # 条件标色：状态非正常 → 红色（优先级最高）
        # 库存量 < 需求量 → 黄色
        if item.get("status", "ok") != "ok":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).fill = red_fill
        elif avail_num is not None and req_num > avail_num:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).fill = yellow_fill

    # --- 工作表 2：已删除 ---
    if deleted_items:
        ws2 = wb.create_sheet("已删除")
        d_headers = ["原行号", "品号", "物料描述", "位号", "数量", "层", "删除原因"]
        ws2.append(d_headers)
        _style_header(ws2, len(d_headers))
        _set_widths(ws2, [8, 14, 45, 40, 8, 8, 20])
        for d in deleted_items:
            ws2.append([
                d.get("original_row", d.get("row", "")),
                d.get("part_number", ""),
                d.get("description", ""),
                d.get("designators", d.get("designator", "")),
                d.get("quantity", ""),
                d.get("layer", ""),
                d.get("reason", ""),
            ])

    # --- 工作表 3：已替换 ---
    if replaced_items:
        ws3 = wb.create_sheet("已替换")
        r_headers = ["原品号", "原描述", "原位号", "需求数量",
                     "替换品号", "替换描述", "仓库", "批次", "库存量", "替换原因"]
        ws3.append(r_headers)
        _style_header(ws3, len(r_headers))
        _set_widths(ws3, [14, 40, 30, 8, 14, 40, 8, 16, 12, 20])
        for r in replaced_items:
            ws3.append([
                r.get("original_part_number", r.get("original_pn", "")),
                r.get("original_description", r.get("original_desc", "")),
                r.get("original_designators", ""),
                r.get("required_quantity", ""),
                r.get("replacement_part_number", r.get("replacement_pn", "")),
                r.get("replacement_description", r.get("replacement_desc", "")),
                r.get("warehouse", r.get("selected_warehouse", "")),
                r.get("batch", r.get("selected_batch", "")),
                r.get("available_quantity", ""),
                r.get("reason", ""),
            ])

    # 统一所有工作表字体
    for ws_to_style in wb.worksheets:
        _apply_font(ws_to_style)

    out_path = input_path.rsplit(".", 1)[0] + "_领料单.xlsx"
    wb.save(out_path)
    print(f"领料单已保存到: {out_path}")
    print(f"  领料单: {len(items)} 条")
    if deleted_items:
        print(f"  已删除: {len(deleted_items)} 条")
    if replaced_items:
        print(f"  已替换: {len(replaced_items)} 条")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python check_bom.py <analyze|fill|delete|extract_parts|gen_picking_list> <参数...>")
        sys.exit(1)

    cmd = sys.argv[1]
    if cmd == "analyze":
        analyze(sys.argv[2])
    elif cmd == "fill":
        # python check_bom.py fill <file> <row> <col_name> <value>
        fill_cell(sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5])
    elif cmd == "delete":
        # python check_bom.py delete <file> <row1> [row2] ...
        delete_rows(sys.argv[2], sys.argv[3:])
    elif cmd == "extract_parts":
        # python check_bom.py extract_parts <file>
        extract_parts(sys.argv[2])
    elif cmd == "gen_picking_list":
        # python check_bom.py gen_picking_list <json_file>
        gen_picking_list(sys.argv[2])
    else:
        print(f"未知命令: {cmd}")
        sys.exit(1)
